import os
import json
import uuid
from datetime import datetime
import streamlit as st
from pathlib import Path
from dotenv import load_dotenv
import openpyxl
import anthropic
import gspread
from google.oauth2.service_account import Credentials

load_dotenv()

EXCEL_PATH = Path(__file__).parent / "quimica_ucv_base_datos.xlsx"
TIMESTAMPS_PATH = Path(__file__).parent / "video_timestamps.json"
CREDENTIALS_PATH = Path(__file__).parent / "google_credentials.json"
SHEET_ID = "1FMy-gOlDEBSErZrk4AAlPiEQWgFzl8bbYcoipSaA1f4"


def get_tracking_sheet():
    try:
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        if CREDENTIALS_PATH.exists():
            creds = Credentials.from_service_account_file(str(CREDENTIALS_PATH), scopes=scopes)
        elif "gcp_service_account" in st.secrets:
            creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=scopes)
        else:
            return None
        gc = gspread.authorize(creds)
        return gc.open_by_key(SHEET_ID).sheet1
    except Exception:
        return None


def log_interaction(sheet, question, answer, session_id):
    if not sheet:
        return
    try:
        urls = [w for w in answer.split() if "idappi.com" in w]
        urls_str = " | ".join(urls[:5]) if urls else ""
        temas = []
        for line in answer.split("\n"):
            if "**" in line and "idappi" not in line:
                tema = line.replace("**", "").strip().lstrip("0123456789. ")
                if tema:
                    temas.append(tema)
        temas_str = " | ".join(temas[:3]) if temas else ""
        sheet.append_row([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            question,
            answer[:500],
            temas_str,
            urls_str,
            session_id,
        ])
    except Exception:
        pass

st.set_page_config(
    page_title="idappi — Asistente Académico",
    page_icon="⚗️",
    layout="centered",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

    /* Reset Streamlit defaults */
    .stApp {
        background: #09090b;
        font-family: 'Inter', sans-serif;
    }
    header[data-testid="stHeader"] { background: transparent; }
    .block-container { padding-top: 2rem; max-width: 800px; }

    /* Hide Streamlit branding and deploy button */
    #MainMenu, footer, .stDeployButton,
    [data-testid="stToolbar"],
    .stAppDeployButton,
    button[kind="header"],
    header[data-testid="stHeader"] button { display: none !important; }

    /* Custom header */
    .app-header {
        display: flex;
        align-items: center;
        gap: 14px;
        padding: 20px 0;
        border-bottom: 1px solid #1c1c1f;
        margin-bottom: 32px;
    }
    .app-logo {
        width: 44px;
        height: 44px;
        background: #d4a017;
        border-radius: 10px;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 22px;
        flex-shrink: 0;
    }
    .app-title {
        font-size: 18px;
        font-weight: 600;
        color: #fafafa;
        letter-spacing: -0.3px;
    }
    .app-subtitle {
        font-size: 13px;
        color: #71717a;
        margin-top: 2px;
    }
    .app-badge {
        margin-left: auto;
        font-size: 11px;
        color: #a1a1aa;
        background: #18181b;
        padding: 4px 10px;
        border-radius: 6px;
        border: 1px solid #27272a;
    }

    /* Stats row */
    .stats-row {
        display: flex;
        gap: 12px;
        margin-bottom: 32px;
    }
    .stat-card {
        flex: 1;
        background: #18181b;
        border: 1px solid #27272a;
        border-radius: 10px;
        padding: 16px 18px;
    }
    .stat-label {
        font-size: 11px;
        color: #71717a;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        margin-bottom: 4px;
    }
    .stat-value {
        font-size: 24px;
        font-weight: 700;
        color: #fafafa;
    }

    /* Chat messages */
    .stChatMessage {
        background: transparent !important;
        border: none !important;
        padding: 12px 0 !important;
    }
    [data-testid="stChatMessageContent"] {
        background: #18181b !important;
        border: 1px solid #27272a !important;
        border-radius: 12px !important;
        padding: 16px 20px !important;
        color: #e4e4e7 !important;
        font-size: 14.5px !important;
        line-height: 1.7 !important;
    }
    [data-testid="stChatMessageContent"] p {
        color: #e4e4e7 !important;
    }
    [data-testid="stChatMessageContent"] strong {
        color: #fafafa !important;
    }
    [data-testid="stChatMessageContent"] a {
        color: #d4a017 !important;
        text-decoration: none !important;
        font-weight: 500 !important;
    }
    [data-testid="stChatMessageContent"] a:hover {
        text-decoration: underline !important;
    }

    /* User message distinction */
    .stChatMessage[data-testid="stChatMessage"]:has([data-testid="chatAvatarIcon-user"]) [data-testid="stChatMessageContent"] {
        background: #1a1a2e !important;
        border-color: #2a2a4a !important;
    }

    /* Chat input — dark box with white text */
    .stChatInput {
        border-top: 1px solid #1c1c1f;
        padding-top: 16px;
    }
    .stChatInput > div,
    .stChatInput [data-testid="stChatInputContainer"],
    [data-testid="stChatInput"] > div,
    [data-testid="stChatInput"] [data-baseweb="textarea"],
    [data-testid="stChatInput"] div[class*="Container"],
    .stChatInput div {
        background: #18181b !important;
        background-color: #18181b !important;
        border-color: #3f3f46 !important;
        border-radius: 12px !important;
    }
    .stChatInput textarea,
    [data-testid="stChatInput"] textarea,
    [data-testid="stChatInputTextArea"] {
        color: #fafafa !important;
        -webkit-text-fill-color: #fafafa !important;
        caret-color: #fafafa !important;
        font-family: 'Inter', sans-serif !important;
        font-size: 14px !important;
        background: transparent !important;
    }
    .stChatInput textarea::placeholder,
    [data-testid="stChatInput"] textarea::placeholder {
        color: #71717a !important;
        -webkit-text-fill-color: #71717a !important;
    }
    /* Send button */
    .stChatInput button,
    [data-testid="stChatInput"] button {
        color: #d4a017 !important;
        background: transparent !important;
    }

    /* Sidebar */
    [data-testid="stSidebar"] {
        background: #0f0f12;
        border-right: 1px solid #1c1c1f;
    }
    [data-testid="stSidebar"] .stMarkdown h3 {
        color: #a1a1aa;
        font-size: 12px;
        text-transform: uppercase;
        letter-spacing: 0.8px;
        margin-bottom: 12px;
    }
    [data-testid="stSidebar"] a {
        color: #a1a1aa !important;
        text-decoration: none !important;
        font-size: 13.5px;
    }
    [data-testid="stSidebar"] a:hover {
        color: #d4a017 !important;
    }
    [data-testid="stSidebar"] .stMarkdown p {
        color: #71717a;
        font-size: 13px;
    }

    /* Welcome card */
    .welcome-card {
        background: #18181b;
        border: 1px solid #27272a;
        border-radius: 14px;
        padding: 32px;
        text-align: center;
        margin: 40px 0;
    }
    .welcome-icon {
        font-size: 36px;
        margin-bottom: 16px;
    }
    .welcome-title {
        font-size: 20px;
        font-weight: 600;
        color: #fafafa;
        margin-bottom: 8px;
    }
    .welcome-desc {
        font-size: 14px;
        color: #71717a;
        line-height: 1.6;
        max-width: 460px;
        margin: 0 auto 24px;
    }
    .suggestions {
        display: flex;
        flex-wrap: wrap;
        gap: 8px;
        justify-content: center;
    }
    .suggestion {
        background: #09090b;
        border: 1px solid #27272a;
        border-radius: 8px;
        padding: 8px 14px;
        font-size: 13px;
        color: #a1a1aa;
        cursor: default;
        transition: border-color 0.2s;
    }
    .suggestion:hover {
        border-color: #d4a017;
        color: #d4a017;
    }

    /* Spinner */
    .stSpinner > div { color: #71717a !important; }

    /* Scrollbar */
    ::-webkit-scrollbar { width: 6px; }
    ::-webkit-scrollbar-track { background: #09090b; }
    ::-webkit-scrollbar-thumb { background: #27272a; border-radius: 3px; }
</style>
""", unsafe_allow_html=True)


@st.cache_data(ttl=3600)
def load_database(_version=3):
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    records = []
    sheets = [
        ("Base de Datos Química UCV", "Química UCV"),
        ("Base de Datos Biología UCV", "Biología UCV"),
        ("Base de Datos Física UCV", "Física UCV"),
    ]
    for sheet_name, curso in sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                break
            records.append({
                "id": row[0], "tipo": row[1], "leccion": row[2],
                "tema": row[3] or "", "nombre_completo": row[4],
                "url": row[5], "keywords": row[6], "descripcion": row[7],
                "curso": curso,
            })
    wb.close()
    return records


@st.cache_data
def load_timestamps():
    if TIMESTAMPS_PATH.exists():
        return json.load(open(TIMESTAMPS_PATH, encoding="utf-8"))
    return {}


@st.cache_data
def build_db_context(_records, _timestamps_json):
    timestamps = json.loads(_timestamps_json) if _timestamps_json else {}
    lines = []
    for r in _records:
        curso = r.get('curso', '')
        lines.append(f"[{r['tipo']}] [{curso}] {r['nombre_completo']}")
        lines.append(f"  Keywords: {r['keywords']}")
        lines.append(f"  Desc: {r['descripcion']}")
        lines.append(f"  URL: {r['url']}")
        lines.append("")

    if timestamps:
        lines.append("\n=== TIMESTAMPS DE VIDEOS ===")
        lines.append("Cada video tiene subtemas con el minuto exacto donde se explican:\n")
        for vid, data in timestamps.items():
            lines.append(f"Video: {data['title']}")
            for t in data['timestamps']:
                lines.append(f"  {t['tiempo']} → {t['subtema']}")
            lines.append("")

    return "\n".join(lines)


SYSTEM_PROMPT = """Eres el asistente académico de idappi.com, plataforma educativa.
Ayudás a los alumnos a encontrar el contenido exacto que necesitan en los cursos disponibles.

CURSOS DISPONIBLES:
- Química UCV (química inorgánica y orgánica)
- Biología UCV (biología celular y molecular)
- Física UCV (magnitudes, vectores, cinemática, dinámica, hidrostática, óptica)

REGLAS:
1. Respondé en español, tono profesional pero cercano.
2. Cuando el alumno pregunte por un tema:
   - Indicá a qué curso pertenece (Química o Biología)
   - Nombrá el tema/lección
   - Incluí el link en formato markdown: [Nombre del tema](URL)
   - Explicá brevemente qué va a encontrar (1 línea)
   - Si tenés datos de timestamps, indicá en qué minuto del video se explica ese subtema específico (ej: "A partir del minuto 03:05")
3. Si hay varios temas relacionados, listá hasta 3, ordenados por relevancia.
4. Si la pregunta no se relaciona con los cursos disponibles, indicalo amablemente.
5. Sé conciso. Tu rol es dirigir al contenido, no explicar el tema.
6. Si el alumno saluda, respondé breve y preguntá en qué tema necesita ayuda.
7. No uses emojis excesivos. Máximo 1-2 por respuesta para mantener profesionalismo.
8. Usá formato limpio con saltos de línea entre cada resultado.
"""


records = load_database()
timestamps_data = load_timestamps()
db_context = build_db_context(tuple(records), json.dumps(timestamps_data))

lecciones = sum(1 for r in records if r['tipo'] == 'Lección')
temas_count = sum(1 for r in records if r['tipo'] == 'Tema')
quizzes = sum(1 for r in records if r['tipo'] == 'Cuestionario')

# Header
st.markdown(f"""
<div class="app-header">
    <div class="app-logo">⚗️</div>
    <div>
        <div class="app-title">Asistente Académico</div>
        <div class="app-subtitle">Química · Biología · Física — idappi.com</div>
    </div>
    <div class="app-badge">v1.0</div>
</div>
""", unsafe_allow_html=True)


# Session state
if "messages" not in st.session_state:
    st.session_state.messages = []
if "session_id" not in st.session_state:
    st.session_state.session_id = str(uuid.uuid4())[:8]
if "tracking_sheet" not in st.session_state:
    st.session_state.tracking_sheet = get_tracking_sheet()

# Welcome card (only if no messages)
if not st.session_state.messages:
    st.markdown("""
    <div class="welcome-card">
        <div class="welcome-icon">⚗️</div>
        <div class="welcome-title">¿En qué tema necesitás ayuda?</div>
        <div class="welcome-desc">
            Escribí tu pregunta y te voy a indicar exactamente dónde encontrar
            el contenido en el curso, con el link directo.
        </div>
        <div class="suggestions">
            <div class="suggestion">Molaridad</div>
            <div class="suggestion">Estructura de Lewis</div>
            <div class="suggestion">Mitosis y Meiosis</div>
            <div class="suggestion">Leyes de Newton</div>
            <div class="suggestion">Caída Libre</div>
            <div class="suggestion">Presión hidrostática</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

# Render messages
for msg in st.session_state.messages:
    with st.chat_message(msg["role"]):
        st.markdown(msg["content"])

# Chat input
if prompt := st.chat_input("Escribí tu pregunta sobre Química..."):
    st.session_state.messages.append({"role": "user", "content": prompt})
    with st.chat_message("user"):
        st.markdown(prompt)

    with st.chat_message("assistant"):
        with st.spinner("Buscando..."):
            try:
                client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
                api_messages = [
                    {"role": m["role"], "content": m["content"]}
                    for m in st.session_state.messages
                ]
                response = client.messages.create(
                    model="claude-sonnet-4-6",
                    max_tokens=1024,
                    system=f"{SYSTEM_PROMPT}\n\nBASE DE DATOS DEL CURSO:\n{db_context}",
                    messages=api_messages,
                )
                answer = response.content[0].text
                st.markdown(answer)
                st.session_state.messages.append({"role": "assistant", "content": answer})
                log_interaction(
                    st.session_state.tracking_sheet,
                    prompt, answer,
                    st.session_state.session_id,
                )
            except Exception as e:
                st.error(f"Error al conectar con el asistente: {e}")

# Sidebar
with st.sidebar:
    st.markdown("### Química UCV")
    for r in records:
        if r['tipo'] == 'Lección' and r.get('curso') == 'Química UCV':
            st.markdown(f"[{r['leccion']}]({r['url']})")
    st.markdown("---")
    st.markdown("### Biología UCV")
    for r in records:
        if r['tipo'] == 'Lección' and r.get('curso') == 'Biología UCV':
            st.markdown(f"[{r['leccion']}]({r['url']})")
    st.markdown("---")
    st.markdown("### Física UCV")
    for r in records:
        if r['tipo'] == 'Lección' and r.get('curso') == 'Física UCV':
            st.markdown(f"[{r['leccion']}]({r['url']})")
    st.markdown("---")
    st.caption("idappi.com — Plataforma educativa")
